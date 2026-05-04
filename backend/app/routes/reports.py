from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_, extract
from sqlalchemy.orm import selectinload
from app.core.database import get_db
from app.core.security import get_current_user, require_roles
from app.models.user import Employee, AttendanceLog, LeaveRequest, LeaveType, StatusEnum, LeaveStatusEnum
from app.models.settings import Holiday
from typing import Optional
from datetime import date, datetime, timedelta
import calendar
import io

router = APIRouter(prefix="/api/reports", tags=["reports"])


def get_working_days(year: int, month: int, holidays: list, working_days_per_week: int = 5) -> int:
    """Count working days in a month excluding weekends and holidays"""
    _, days_in_month = calendar.monthrange(year, month)
    holiday_dates = {h.date for h in holidays}
    working = 0
    max_weekday = 4 if working_days_per_week == 5 else 5  # Mon-Fri or Mon-Sat
    for day in range(1, days_in_month + 1):
        d = date(year, month, day)
        if d.weekday() <= max_weekday and d not in holiday_dates:
            working += 1
    return working


@router.get("/monthly-summary")
async def monthly_summary(
    year: int = Query(default=None),
    month: int = Query(default=None),
    department_id: Optional[int] = None,
    shift_id: Optional[int] = None,
    employee_id: Optional[int] = None,
    office_id: Optional[int] = None,
    db: AsyncSession = Depends(get_db),
    _=Depends(require_roles("admin", "hr", "manager"))
):
    today = date.today()
    year = year or today.year
    month = month or today.month

    start_date = date(year, month, 1)
    _, last_day = calendar.monthrange(year, month)
    end_date = date(year, month, last_day)

    # Get holidays for this month
    holidays_result = await db.execute(
        select(Holiday).where(
            and_(Holiday.date >= start_date, Holiday.date <= end_date)
        )
    )
    holidays = holidays_result.scalars().all()
    from app.models.settings import CompanySettings as CS
    cs_result = await db.execute(select(CS).limit(1))
    cs = cs_result.scalar_one_or_none()
    working_days_per_week = cs.working_days_per_week if cs and hasattr(cs, 'working_days_per_week') else 5
    working_days = get_working_days(year, month, holidays, working_days_per_week)

    # Get employees
    from app.models.settings import EmployeeOffice
    emp_query = select(Employee).options(
        selectinload(Employee.department),
        selectinload(Employee.shift)
    ).where(Employee.is_active == True)
    if department_id:
        emp_query = emp_query.where(Employee.department_id == department_id)
    if shift_id:
        emp_query = emp_query.where(Employee.shift_id == shift_id)
    if employee_id:
        emp_query = emp_query.where(Employee.id == employee_id)
    if office_id:
        office_emp_ids = (await db.execute(
            select(EmployeeOffice.employee_id).where(EmployeeOffice.office_id == office_id)
        )).scalars().all()
        emp_query = emp_query.where(Employee.id.in_(office_emp_ids))
    employees = (await db.execute(emp_query)).scalars().all()

    summary = []
    for emp in employees:
        # Get attendance logs
        logs_result = await db.execute(
            select(AttendanceLog).where(
                and_(
                    AttendanceLog.employee_id == emp.id,
                    AttendanceLog.date >= start_date,
                    AttendanceLog.date <= end_date
                )
            )
        )
        logs = logs_result.scalars().all()

        present = sum(1 for l in logs if l.status == StatusEnum.present)
        late = sum(1 for l in logs if l.status == StatusEnum.late)
        on_leave = sum(1 for l in logs if l.status == StatusEnum.on_leave)
        absent = working_days - present - late - on_leave
        early_leaves = sum(1 for l in logs if getattr(l, 'early_leave', False))

        # Calculate total hours
        total_minutes = 0
        for log in logs:
            if log.check_in_time and log.check_out_time:
                diff = log.check_out_time - log.check_in_time
                total_minutes += diff.total_seconds() / 60
        total_hours = round(total_minutes / 60, 1)

        # Leave balance
        leave_used = await db.execute(
            select(func.count(LeaveRequest.id)).where(
                and_(
                    LeaveRequest.employee_id == emp.id,
                    LeaveRequest.status == LeaveStatusEnum.approved,
                    extract('year', LeaveRequest.from_date) == year
                )
            )
        )
        leaves_taken = leave_used.scalar() or 0

        summary.append({
            "employee_id": emp.id,
            "employee_code": emp.employee_code,
            "full_name": emp.full_name,
            "department": emp.department.name if emp.department else "-",
            "shift": emp.shift.name if emp.shift else "-",
            "working_days": working_days,
            "present": present,
            "late": late,
            "absent": max(0, absent),
            "on_leave": on_leave,
            "total_hours": total_hours,
            "leaves_taken": leaves_taken,
            "attendance_pct": round((present + late) / working_days * 100, 1) if working_days > 0 else 0,
            "early_leaves": early_leaves,
        })

    return {
        "year": year,
        "month": month,
        "month_name": calendar.month_name[month],
        "working_days": working_days,
        "holidays": [{"date": str(h.date), "name": h.name} for h in holidays],
        "summary": sorted(summary, key=lambda x: x["full_name"])
    }


@router.get("/late-report")
async def late_report(
    year: int = Query(default=None),
    month: int = Query(default=None),
    department_id: Optional[int] = None,
    shift_id: Optional[int] = None,
    employee_id: Optional[int] = None,
    db: AsyncSession = Depends(get_db),
    _=Depends(require_roles("admin", "hr", "manager"))
):
    today = date.today()
    year = year or today.year
    month = month or today.month
    start_date = date(year, month, 1)
    _, last_day = calendar.monthrange(year, month)
    end_date = date(year, month, last_day)

    query = (
        select(AttendanceLog)
        .options(
            selectinload(AttendanceLog.employee).selectinload(Employee.department),
            selectinload(AttendanceLog.employee).selectinload(Employee.shift)
        )
        .where(
            and_(
                AttendanceLog.status == StatusEnum.late,
                AttendanceLog.date >= start_date,
                AttendanceLog.date <= end_date
            )
        )
    )
    if employee_id:
        query = query.where(AttendanceLog.employee_id == employee_id)
    if department_id:
        query = query.join(Employee, AttendanceLog.employee_id == Employee.id).where(Employee.department_id == department_id)
    if shift_id:
        query = query.join(Employee, AttendanceLog.employee_id == Employee.id).where(Employee.shift_id == shift_id)
    query = query.order_by(AttendanceLog.date.desc())
    result = await db.execute(query)
    logs = result.scalars().all()

    return [{
        "date": str(log.date),
        "employee_code": log.employee.employee_code,
        "full_name": log.employee.full_name,
        "department": log.employee.department.name if log.employee.department else "-",
        "check_in_time": log.check_in_time.strftime("%H:%M:%S") if log.check_in_time else "-",
    } for log in logs]


@router.get("/absent-report")
async def absent_report(
    from_date: Optional[date] = None,
    to_date: Optional[date] = None,
    employee_id: Optional[int] = None,
    department_id: Optional[int] = None,
    shift_id: Optional[int] = None,
    office_id: Optional[int] = None,
    db: AsyncSession = Depends(get_db),
    _=Depends(require_roles("admin", "hr", "manager"))
):
    today = date.today()
    from_date = from_date or date(today.year, today.month, 1)
    to_date = to_date or today

    from app.models.settings import EmployeeOffice
    emp_query = select(Employee).options(
        selectinload(Employee.department),
        selectinload(Employee.shift)
    ).where(Employee.is_active == True)
    if employee_id:
        emp_query = emp_query.where(Employee.id == employee_id)
    if department_id:
        emp_query = emp_query.where(Employee.department_id == department_id)
    if shift_id:
        emp_query = emp_query.where(Employee.shift_id == shift_id)
    if office_id:
        office_emp_ids = (await db.execute(
            select(EmployeeOffice.employee_id).where(EmployeeOffice.office_id == office_id)
        )).scalars().all()
        emp_query = emp_query.where(Employee.id.in_(office_emp_ids))
    employees = (await db.execute(emp_query)).scalars().all()

    holidays_result = await db.execute(
        select(Holiday).where(and_(Holiday.date >= from_date, Holiday.date <= to_date))
    )
    holiday_dates = {h.date for h in holidays_result.scalars().all()}

    report = []
    current = from_date
    while current <= to_date:
        if current.weekday() < 5 and current not in holiday_dates:
            for emp in employees:
                log = (await db.execute(
                    select(AttendanceLog).where(
                        and_(AttendanceLog.employee_id == emp.id, AttendanceLog.date == current)
                    )
                )).scalar_one_or_none()
                if not log:
                    report.append({
                        "date": str(current),
                        "employee_code": emp.employee_code,
                        "full_name": emp.full_name,
                        "department": emp.department.name if emp.department else "-",
                        "shift": emp.shift.name if emp.shift else "-",
                    })
        current += timedelta(days=1)

    return sorted(report, key=lambda x: (x["date"], x["full_name"]))


@router.get("/export/excel")
async def export_excel(
    year: int = Query(default=None),
    month: int = Query(default=None),
    db: AsyncSession = Depends(get_db),
    _=Depends(require_roles("admin", "hr"))
):
    today = date.today()
    year = year or today.year
    month = month or today.month

    # Get summary data
    summary_data = await monthly_summary(year=year, month=month, department_id=None, db=db, _=None)

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return {"error": "openpyxl not installed"}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{calendar.month_name[month]} {year}"

    # Header styles
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Title
    ws.merge_cells('A1:K1')
    title_cell = ws['A1']
    title_cell.value = f"Monthly Attendance Report — {calendar.month_name[month]} {year}"
    title_cell.font = Font(bold=True, size=14, color="1E40AF")
    title_cell.alignment = center
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:K2')
    ws['A2'].value = f"Working Days: {summary_data['working_days']} | Generated: {datetime.now().strftime('%d %b %Y %H:%M')}"
    ws['A2'].alignment = center
    ws['A2'].font = Font(italic=True, color="6B7280")
    ws.row_dimensions[2].height = 20

    # Column headers
    headers = ['#', 'Code', 'Employee Name', 'Department', 'Shift', 'Present', 'Late', 'Absent', 'On Leave', 'Total Hours', 'Attendance %']
    ws.row_dimensions[4].height = 22
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    # Data rows
    green_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    yellow_fill = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")

    for row_idx, emp in enumerate(summary_data['summary'], 1):
        row = row_idx + 4
        ws.row_dimensions[row].height = 18
        values = [row_idx, emp['employee_code'], emp['full_name'], emp['department'], emp['shift'],
                  emp['present'], emp['late'], emp['absent'], emp['on_leave'], emp['total_hours'], f"{emp['attendance_pct']}%"]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = border
            cell.alignment = center if col != 3 else Alignment(horizontal="left", vertical="center")
            # Color attendance %
            if col == 11:
                pct = emp['attendance_pct']
                if pct >= 90: cell.fill = green_fill
                elif pct >= 75: cell.fill = yellow_fill
                else: cell.fill = red_fill

    # Column widths
    widths = [4, 8, 22, 16, 14, 8, 7, 8, 9, 12, 12]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"attendance_report_{year}_{month:02d}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/my-attendance")
async def my_attendance(
    year: int = Query(default=None),
    month: int = Query(default=None),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user)
):
    """Employee self-service - view own attendance"""
    today = date.today()
    year = year or today.year
    month = month or today.month
    start_date = date(year, month, 1)
    _, last_day = calendar.monthrange(year, month)
    end_date = date(year, month, last_day)

    emp_result = await db.execute(
        select(Employee).options(selectinload(Employee.shift))
        .where(Employee.user_id == current_user.id)
    )
    emp = emp_result.scalar_one_or_none()
    if not emp:
        # Admin viewing - show all employees summary instead
        return {"error": None, "is_admin": True, "message": "Admin account has no employee profile. Please view Reports for full attendance data."}

    logs_result = await db.execute(
        select(AttendanceLog).where(
            and_(
                AttendanceLog.employee_id == emp.id,
                AttendanceLog.date >= start_date,
                AttendanceLog.date <= end_date
            )
        ).order_by(AttendanceLog.date)
    )
    logs = logs_result.scalars().all()

    holidays_result = await db.execute(
        select(Holiday).where(and_(Holiday.date >= start_date, Holiday.date <= end_date))
    )
    holiday_dates = {h.date: h.name for h in holidays_result.scalars().all()}

    # Build calendar data
    calendar_data = []
    current = start_date
    while current <= end_date:
        log = next((l for l in logs if l.date == current), None)
        is_weekend = current.weekday() >= 5
        is_holiday = current in holiday_dates

        if is_weekend:
            status = "weekend"
        elif is_holiday:
            status = "holiday"
        elif log:
            status = log.status.value
        elif current < today:
            status = "absent"
        else:
            status = "future"

        calendar_data.append({
            "date": str(current),
            "day": current.day,
            "weekday": current.strftime("%a"),
            "status": status,
            "holiday_name": holiday_dates.get(current),
            "check_in": log.check_in_time.strftime("%H:%M") if log and log.check_in_time else None,
            "check_out": log.check_out_time.strftime("%H:%M") if log and log.check_out_time else None,
        })
        current += timedelta(days=1)

    # Leave balance
    leave_types = (await db.execute(select(LeaveType))).scalars().all()
    leave_balance = []
    for lt in leave_types:
        used = (await db.execute(
            select(func.count(LeaveRequest.id)).where(
                and_(
                    LeaveRequest.employee_id == emp.id,
                    LeaveRequest.leave_type_id == lt.id,
                    LeaveRequest.status == LeaveStatusEnum.approved,
                    extract('year', LeaveRequest.from_date) == year
                )
            )
        )).scalar() or 0
        leave_balance.append({
            "type": lt.name,
            "allocated": lt.days_per_year,
            "used": used,
            "remaining": lt.days_per_year - used
        })

    present = sum(1 for l in logs if l.status == StatusEnum.present)
    late = sum(1 for l in logs if l.status == StatusEnum.late)
    on_leave = sum(1 for l in logs if l.status == StatusEnum.on_leave)

    return {
        "employee": {"name": emp.full_name, "code": emp.employee_code},
        "year": year, "month": month,
        "month_name": calendar.month_name[month],
        "summary": {"present": present, "late": late, "on_leave": on_leave},
        "calendar": calendar_data,
        "leave_balance": leave_balance
    }
